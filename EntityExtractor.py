from voice import Speak
from Speech import *
import re
import openpyxl as xl
import random
import en_core_web_sm
import spacy
import datetime
from nltk import word_tokenize, pos_tag
import Config
ent_func = None
default_responses = ["I did not understand ", "Can you repeate it", "I still can't get it"]
Doc = en_core_web_sm.load()
responses = ""

def speakDefaultResponse():
    global default_responses
    Speak(random.choice(default_responses))
def isCancelled(tx):
    if tx.lower().find("cancel") != -1 or tx.lower().find('exit') != -1:
        return "CANCELED"


class Extractor():
    def __init__(self, ent_list, text):
        self.ent_list = ent_list
        self.text = text
        self.extracted_ents = {}  # Stores ->  { "TIME":["2:34 pm","12:22 am"] }

    def getSpecifiedEntites(self):
        for i in self.ent_list:
            feedback = self.extract(i)
            if feedback=="CANCELLED":
                return "CANCELLED"
        return self.extracted_ents

    def extract(self, ent):
        func = None
        tx = self.text
        e = []
        if ent['name'] in ent_func:
            func = ent_func[ent['name']]

        else:
            return "NOT_FOUND_ENTITY"
        first_time = True
        while True:
            responses = ent['response']
            output = func(tx)
            # Ask for Entity
            if output == [] and ent['required'] == 1:
                # Default feedback
                if not first_time:
                    speakDefaultResponse()
                else:
                    first_time = False
                if type(ent['response'])==list:
                    Speak(random.choice(ent['response']))
                else:
                    Speak(ent['response'])
                tx = listen()  # Change required
                if isCancelled(tx) :
                    return "CANCELLED"
            else:
                if output=="CANCELLED":
                    return "CANCELLED"
                e.extend(output)
                self.extracted_ents[ent['name']] = e
                return "FOUND"
    

def isContainAnyDay(day):
    days = ["sunday", "monday", "tuesday", "wednesday", "thursday", "friday", "saturday"]
    for i in days:
        if i == day:
            return True
    return False

def extractTime(text):
    l = [None,None]
    text = text.lower()
    #  Extract Time --------------------------------
    from tsresolve import point_of_time,period_of_time
    t = point_of_time(text)
    if t[0] != None:
        l[0]= datetime.datetime.strptime(t[0][-8:],"%H:%M:%S")
    # * Regex
    res = re.findall(r'\d+\s+[pa][m]',text)
    if res!=[]:
        # 4 pm
        if int(res[0][:2]) <= 12:
            l[0]=datetime.datetime.strptime(res[0],"%I %p")
    res = re.findall(r'((([0]?[1-9]|1[0-2])(:|\.)[0-5][0-9]((:|\.)[0-5][0-9])?( )?(AM|am|aM|Am|PM|pm|pM|Pm))|(([0]?[0-9]|1[0-9]|2[0-3])(:|\.)[0-5][0-9]((:|\.)[0-5][0-9])?))',text)
    if res!=[]:
        # 3:04 am
        if res[0][0][-2:]=='pm' or res[0][0][-2:]=='am' :
            t = datetime.datetime.strptime(res[0][0],"%I:%M %p")
            l[0]=t
    
    
    
    
    #  Time Duration --------------------------------
    l[1] = time_duration(text)

    if l[0]==None and l[1]==None:
        return []
    return l
def time_duration(text):
    from datetime import datetime
    text_words = text.lower().split(" ")
    numbers = re.findall(r'\d+',text)
    if numbers==[]:
        return None
    words = ['sec','seconds','second','min','mins','minutes','minutes','hr','hrs','hour','hours']
    t = []
    for ws in text_words :
        if ws in words:
            t.append(ws[0])
    if t==[]:
        return None
    # Create Time
    h,m,s = 0,0,0
    for index in range(len(numbers)):
        no = int(numbers[index])
        if t[index]=='h':
            h = no
        if t[index]=='m':
            m = no
        if t[index]=='s':
            s = no
    return datetime(1,1,1,h,m,s)
def extractDate(text):
    global Doc
    doc = Doc(text)
    l = []
    import datefinder
    for i in datefinder.find_dates(text):
        l.append(i)
    ent_list = [(X.text, X.label_) for X in doc.ents]    
    print(ent_list,l)
    try:
        for i in ent_list:
            if i[1] == "DATE":
                for i in datefinder.find_dates(i[0]):
                    l.append(i)
                if i[0] == "today":
                    today = datetime.datetime.today()
                    l.append(today)

                if isContainAnyDay(i[0]):
                    today = datetime.date.today()
                    if i[0] == "monday":
                        day = today + datetime.timedelta((0 - today.weekday()) % 7)
                    elif i[0] == "tuesday":
                        day = today + datetime.timedelta((1 - today.weekday()) % 7)
                    elif i[0] == "wednesday":
                        day = today + datetime.timedelta((2 - today.weekday()) % 7)
                    elif i[0] == "thursday":
                        day = today + datetime.timedelta((3 - today.weekday()) % 7)
                    elif i[0] == "friday":
                        day = today + datetime.timedelta((4 - today.weekday()) % 7)
                    elif i[0] == "saturday":
                        day = today + datetime.timedelta((5 - today.weekday()) % 7)
                    elif i[0] == "sunday":
                        day = today + datetime.timedelta((6 - today.weekday()) % 7)
                    l.append(day)
                else:
                    print(i[0])
    except:
        pass
    return l

def extractPerson(text):
    global Doc
    doc = Doc(text)
    ent_list = [(X.text, X.label_) for X in doc.ents]
    print(ent_list)
    l = []
    for i in ent_list:
        if i[1] == "PERSON":
            l.append(i[0])
    return l

def extractLocation(text):
    global Doc
    doc = Doc(text)
    ent_list = [(X.text, X.label_) for X in doc.ents]
    print(ent_list)
    l = []
    for i in ent_list:
        if i[1] == "LOCATION":
            l.append(i[0])
        if i[1] == "GPE":
            l.append(i[0])
    return l

def showEntities(text):
    global Doc
    doc = Doc(text)
    ent_list = [(X.text, X.label_) for X in doc.ents]
    print(ent_list)

def extractMoney(text):
    global Doc
    doc = Doc(text)
    ent_list = [(X.text, X.label_) for X in doc.ents]
    print(ent_list)
    l = []
    for i in ent_list:
        if i[1] == "MONEY":
            l.append([0])
    return l

def extractNumber(text):
    global Doc
    doc = Doc(text)
    ent_list = [(X.text, X.label_) for X in doc.ents]
    print(ent_list)
    l = []
    for i in ent_list:
        if i[1] == "NUMBER":
            l.append([0])
    return l
def doMaths(text):
    import math
    op,ans = None,None
    # Initialising
    square = lambda x : x*x
    cube = lambda x : x*x*x 
    operation = {
        "+,plus,add":"+",
        "-,minus,subtract":"-",
        "*,times,multiply,into":"*",
        "/,divide":"/",
        "raise,raise to,^,power":"**",
        'factorial,!':math.factorial,
        'square':square,
        "cube":cube,
        'square root':math.sqrt,
        'tan':math.tan,
        'sin':math.sin,
        'cos':math.cos 
    }
    # Find operation to be done
    for i in operation:
        k = i.split(",")
        for j in k:
            if j in text.lower():
                op=operation[i]
                break 
    if op==None: 
        Speak("I did'nt found any operation")
        return []
    numbers = re.findall(r'\d+',text)
    if numbers == []:
        Speak("Sir the operands are missing")
        return []
    # Binary Operation
    if isinstance(op,str):
        if len(numbers) < 2:
            Speak("Sir the operands are missing")
        print("exp:"+numbers[0]+op+numbers[1])
        ans = eval(numbers[0]+op+numbers[1])
    # Functions
    else:
        print("exp:",op,numbers[0])
        ans = op(int(numbers[0]))
    print("Answer=",ans)
    return [str(ans)]
def extractMail(text):
    l = []
    # If person specified in text ->  Find in Contacts
    # contacts = Config.getContacts()
    
    # If email specified in text -> extract email from text
    mail = re.findall(r'[\w\.-]+@[\w\.-]+', text)
    if mail != [] :
         l.extend(mail)
         return l
    
    wb = xl.load_workbook('Data/Contacts.xlsx')
    sheet = wb.active
    max_cols = sheet.max_column
    max_row = sheet.max_row
    contact_list = []
    for i in range(1,sheet.max_row+1):
        if compare_lists(text.lower().split(" "),sheet['A'+str(i)].value.lower().split(' ')) > 0 :
            contact = []
            for j in range(1,sheet.max_column+1):
                contact.append(sheet[chr(64+j)+str(i)].value)
            contact_list.append(contact)
    print(contact_list)
    if len(contact_list) == 0 :
        Speak("I didn't found anyone with such name")
        return l
    elif len(contact_list) > 1 :
        Speak("Sir i found multiple records with similar names")
        return []
    elif contact_list[0][1]==None:
        Speak("Sir you didn't specfied  email address for this person. You can enter it manually")
    else:
        contact_list[0][2:] = []
        contact_list[0].reverse()
        l.extend(contact_list[0])
        return l
    
    
def extractApp_and_Tag(text):
    l = []
    text_tok = lower_text(word_tokenize(text))
    apps_dict = Config.getApps()
    tag_dict = Config.get("TAGS")
    score_app,score_tag = 0,0
    found_app,found_tag = None,None
    # Searching for Apps
    for app in apps_dict:
        app_tok = lower_text(word_tokenize(app))
        if(compare_lists(text_tok,app_tok)>score_app):
            score_app = compare_lists(text_tok,app_tok)
            found_app = app
    # Searching for tags
    for tag in tag_dict:
        tag_tok = lower_text(word_tokenize(tag))
        if(compare_lists(text_tok,tag_tok)>score_tag):
            score_tag = compare_lists(text_tok,tag_tok)
            found_tag = tag
    # Nothing Found
    if score_app==0 and score_tag==0:
        return l
    # Check who matched the best App or Tag    
    if score_tag>=score_app :
        l.append(found_tag)
        l.append(tag_dict[found_tag])
        l.append("TAG")
    else:
        l.append(found_app)
        l.append(apps_dict[found_app])
        l.append("APP")
        
    return l

def lower_text(text_list):
    for i in range(len(text_list)):
        text_list[i]=text_list[i].lower()
    return text_list
def compare_lists(l1,l2):
    score = 0
    for i in l1:
        if i in l2:
            score+=1

    return score

def extractMessage(text):
    l = []
    Speak("What is the message?")
    msg  = listen()
    if msg==None:
        return l
    if isCancelled(msg):
        return "CANCELLED"
    l.append(msg)
    return l
    

def extractValue(text):
    l = []
    v = ""

    if text.find("increase")!=-1:
        v+="+"
    elif text.find("decrease")!=-1 or text.find("reduce")!=-1:
        v+="-"
    else:
        v+="="
    no = re.findall("\d{1,3}",text)
    if no!=[]:
        v+=no[0]
    if v=="=":
        return []
    l.append(v)
    return l
def extractText(text):
    return [text]

def extractEvent(text):

    Speak("What is the event?")
    msg = listen()
    if msg==None:
        return []
    if isCancelled(msg):
        return "CANCELLED"
    return [msg] 


    # text_token = lower_text(word_tokenize(text))
    # event_list =['birthday','anniversary','event','holiday',]
    # event_found = None
    # for i in text_token:
    #     if i in event_list:
    #         l.append(i)
    # if len(l)==0:
    #     return []
    # # exrtact event Description
    # for i in ['as','that','of']:
    #     if i in text_token:
    #         l.append(" ".join(text_token[text_token.index(i)+1:]))
    #         break
    # else:
    #     l.append(" ".join(text_token)) 
    # return l

ent_func = {
            "TIME": extractTime,
            "DATE": extractDate,
            "PERSON": extractPerson,
            "MONEY": extractMoney,
            "LOCATION": extractLocation,
            "APP|TAG": extractApp_and_Tag,
            "MESSAGE": extractMessage,
            "EMAIL_ADDRESS": extractMail,
            "NUMBER": extractNumber,
            "VALUE":extractValue,
            "TEXT":extractText,
            "EVENT":extractEvent,
            "MATHS":doMaths
            }

def extractAllEnts(text):
    ent_list = {}
    for i in ent_func:
        arr = ent_func[i](text)
        ent_list.update({i: arr})
    print(ent_list)
# ENTITY : FUNCTION list
# def sample(text):
#     l = []
#     mail = re.findall(r'\b((1[0-2]|0?[1-9]):([0-5][0-9]) ([AaPp][Mm]))', text)
#     print(mail[0][0])
#     # l.append(mail)
#     # return l

if __name__ == "__main__":
    print(extractMail("i want to it  Nikhil"))
    # print(extractMsg("Hi to Nikhil"))
    pass

